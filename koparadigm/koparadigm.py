import openpyxl
from jamo import h2j, j2h, hcj_to_jamo, is_hcj
import re
import os
from collections import defaultdict

RESOURCE_PATH = os.path.dirname(os.path.abspath(__file__)) + "/koparadigm.xlsx"
RESOURCE = openpyxl.load_workbook(filename=RESOURCE_PATH, read_only=True)


class Paradigm(object):
    def __init__(self):
        self.verb2verb_classes = self.make_verb2verb_classes()
        self.ending_class2endings = self.make_ending_class2endings()
        self.verb_class2rules = self.make_verb_class2rules()

    def make_verb2verb_classes(self):
        verb2verb_classes = defaultdict(list)  # e.g., {"곱": [1,2]}

        sh = RESOURCE["Verbs"]
        for verb, verb_class in sh.iter_rows(
            min_col=2, max_col=3, min_row=2, values_only=True
        ):
            verb2verb_classes[verb].append(verb_class)
        return verb2verb_classes


    def make_ending_class2endings(self):
        ending_class2endings = defaultdict(list)  # e.g., {1: ["어야", "어서]}

        sh = RESOURCE["Endings"]
        for ending, ending_class in sh.iter_rows(
            min_col=2, max_col=3, min_row=2, values_only=True
        ):
            ending_class2endings[ending_class].append(ending)
        return ending_class2endings


    def make_verb_class2rules(self):
        verb_class2rules = defaultdict(list)
        sh = RESOURCE["Template"]

        ending_classes = list(
            next(sh.iter_rows(min_row=1, max_row=1, min_col=3, values_only=True))
        )

        for row in sh.iter_rows(min_row=3, values_only=True):
            verb_class = row[0]
            for ending_class, rule in zip(ending_classes, row[2:]):
                if rule is not None:
                    rule = rule[1:-1]  # ( ... )
                else:
                    rule = ""
                rule = (ending_class, rule)
                verb_class2rules[verb_class].append(rule)

        return verb_class2rules


    def contract(self, string):
        '''vowel contraction'''
        _string = string
        pairs = [
            ("ᅡ|ᅡ","ᅡ", "required"),
            ("ᅥ|ᅥ","ᅥ", "required"),
            ("오|ᅡ","와", "required"),
            ("ᅩ|ᅡ","ᅪ", "optional"),
            ("ᅮ|ᅥ","ᅯ", "optional"),
            ("ᅳ|ᅥ","ᅥ", "required"),
            ("ᅵ|ᅥ","ᅧ", "optional"),
            ("ᅢ|ᅥ","ᅢ", "optional"),
            ("ᅦ|ᅥ","ᅦ", "optional"),
            ("ᅬ|ᅥ","ᅫ", "optional"),
            ("하|ᅧ","해", "optional"),
        ]

        for pair in pairs:
            untouched, contraction, cond = pair
            if untouched in string:
                string = string.replace(untouched, contraction)
                if cond == "optional":
                    string = _string.replace("|", "ᄋ") + "/" + string
                return string
        return _string.replace("|", "")

    def compose(self, string):
        string = self.contract(string)

        choseong = "[\u1100-\u1112]"
        jungseong = "[\u1161-\u1175]"
        jongseong = "[\u11A8-\u11C2]"

        # CVC first
        matches = re.findall(f"{choseong}{jungseong}{jongseong}", string)
        for match in matches:
            syl = j2h(*match)
            string = string.replace(match, syl)

        # CV
        matches = re.findall(f"{choseong}{jungseong}", string)
        for match in matches:
            syl = j2h(*match)
            string = string.replace(match, syl)

        return string

    def combine(self, verb, ending, rule):
        if not rule:
            return []

        stop, postfix, start = rule.split(",")
        stop = None if stop == "" else int(stop)
        start = None if start == "" else int(start)

        # STEP 1. Decompose verb
        verb = h2j(verb) # h: hangul syl. j: jamo

        # STEP 2. Slice 1
        verb = verb[:stop]

        # STEP 3. Merge 2 and postfix
        wordform = verb + postfix

        # STEP 4. Decompose ending
        ending = h2j(ending)
        ending = "".join(hcj_to_jamo(char, "tail") if is_hcj(char) else char for char in ending)

        # STEP 5. Slice 4
        ending = ending[start:]

        # STEP 6. Merge 3 and 5
        wordform +="|" + ending

        # STEP 7. Compose 6
        wordform = self.compose(wordform)

        return wordform


    def conjugate(self, verb):
        '''
        Main method
        '''
        if verb in self.verb2verb_classes:
            paradigms = []
            verb_classes = self.verb2verb_classes[verb]
            for verb_class in verb_classes:
                # pos
                if verb_class < 3:
                    pos = "Action Verb/Descriptive Verb"
                elif verb_class == 14: # copula
                    pos = "Copula"
                elif (verb_class <= 6) or (15 <= verb_class <= 30):
                    pos = "Action Verb"
                else:
                    pos = "Descriptive Verb"
                # ending info.
                li_ending_and_form = []
                rules = self.verb_class2rules[verb_class]
                for ending_class, rule in rules:
                    endings = self.ending_class2endings[ending_class]
                    for ending in endings:
                        form = self.combine(verb=verb, ending=ending, rule=rule)
                        if len(form) > 0:
                            li_ending_and_form.append((ending, form))

                paradigms.append([pos, li_ending_and_form])
            return paradigms
        else:
            print(f"{verb} is NOT found.")


def prettify(paradigms):
    '''utility function. pretty print.
    '''
    for num, paradigm in enumerate(paradigms, start=1):
        pos, li_ending_and_form = paradigm
        print("="*20, num, "="*20)
        print("POS =", pos)
        for ending, form in li_ending_and_form:
            print("• ending = {} form = {}".format(ending, form))
        print()